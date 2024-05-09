import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

def Crawl_CTBV(url, href):
    data = []  # Create an empty list to store crawled data
    response = requests.get(url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')    
        div = soup.find('div', {'id': 'mainContent'})
        if div: 
            ps = div.find_all('p')
            if ps:
                for p in ps:
                    dt = {}
                    ct = p.text
                    I = p.find('img')
                    img = I.get('src') if I else ""
                    if len(ct) > 0:
                        dt["href"] = href
                        dt["content"] = ct
                        dt["img"] = ""
                        data.append(dt)
                    if len(img) > 0:
                        dt["href"] = href
                        dt["content"] = ""
                        dt["img"] = img
                        data.append(dt)
    else:
        print("huhu")
    return data

def Write_to_Excel(data, file_name):
    wb = Workbook()
    sheet_name = "Chi Tiết Bài Viết"
    worksheet = wb.active
    worksheet.title = sheet_name
    worksheet.append(["Href", "Content", "Image"])
    for item in data:
        worksheet.append([item.get("href", ""), item.get("content", ""), item.get("img", "")])
    wb.save(f'D:/TDMU/Nam3/HK2/KTLTinPTTK/project/Test/{file_name}.xlsx')
    print('Lưu thành công')

def read_excel(file_name):
    data = []
    wb = load_workbook(filename=file_name)
    ws = wb.active
    rows = ws.iter_rows(min_row=2, values_only=True)
    for row in rows:
        topic, amh, title, href, time, descript = row[:6]
        data.append((topic, amh, title, href, time, descript))
    wb.close()
    return data   

if __name__ == '__main__':
    file_name = 'data_TinTuc.xlsx'
    data_TinTuc = read_excel(file_name)
    data_post_detail = []
    for topic, amh, title, href, time, descript in data_TinTuc:
        url = "https://thi.tuyensinh247.com" + href
        da = Crawl_CTBV(url, href)
        data_post_detail.extend(da)
    Write_to_Excel(data_post_detail, 'ChiTietBaiViet')
