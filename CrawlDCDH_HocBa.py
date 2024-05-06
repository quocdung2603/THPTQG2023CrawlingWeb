from openpyxl import Workbook, load_workbook
import requests
from bs4 import BeautifulSoup

worksheet_dict = {}
def Crawl_DCDH_HocBa(url, tenDH, tvt):
    data=[]
    response = requests.get(url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        div = soup.find('div',{'id':'tab_2'})
        if div:
            tables = div.find('table', {'width':'100%','border':'0','cellpadding':'0','cellspacing':'0'})
            if tables:
                trs = tables.find_all('tr',{'class':'bg_white'})
                if trs:
                    for tr in trs:
                        dt = [cell.text for cell in tr.find_all('td')]
                        data.append(dt)
                    Write_to_excel(wb, tvt, data)
    else:
        print("huhu")

def Write_to_excel(wb, sheet_name, data):
    if sheet_name not in worksheet_dict:
        worksheet = wb.active
        worksheet.title = sheet_name
        worksheet_dict[sheet_name] = worksheet
        worksheet.append(["STT", "Mã ngành", "Tên ngành", "Tổ hợp môn", "Điểm chuẩn", "Ghi chú"])
    else: 
        worksheet = worksheet_dict[sheet_name]
    for row in data:
        worksheet.append(row)   
        
    print("Điểm chuẩn của",sheet_name)

def read_excel(file_name):
    data = []
    wb = load_workbook(filename=file_name)
    ws = wb.active
    # Bỏ qua dòng đầu tiên (chứa tiêu đề cột)
    rows = ws.iter_rows(min_row=2, values_only=True)
    for row in rows:
        tenDH, tvt, href = row[:3]  # Truy xuất 2 cột đầu tiên trong mỗi dòng
        data.append((tenDH, tvt, href))
    wb.close()
    return data   

if __name__ == '__main__':
    file_name = 'data_TenDH.xlsx'
    data_DH = read_excel(file_name)
    for tenDH, tvt, href in data_DH:
        wb = Workbook()
        url = "https://diemthi.tuyensinh247.com" + str(href)
        Crawl_DCDH_HocBa(url, tenDH, tvt)
        wb.save(f'D:/TDMU/Nam3/HK2/KTLTinPTTK/project/DCDH2023_HB/{tvt}.xlsx')