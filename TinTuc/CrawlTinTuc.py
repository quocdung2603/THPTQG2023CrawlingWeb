import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import pyodbc
conx = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=MSI\\SQLEXPRESS;DATABASE=Diem_Thi_THPTQG;UID=lekiet;PWD=123456')
cursor = conx.cursor()
data=[]
def Crawl_News(url, topic):
    response = requests.get(url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content,'html.parser')    
        ct=soup.find('div', {'class': 'box-cont clearfix'})
        uls = ct.find('ul', {'class': 'list clearfix'})
        
        if uls:
            lis = uls.find_all('li',{'class':'clearfix'})
            if lis:
                for li in lis: 
                    Img = li.find('img',{'class':'img-132'})
                    Img = Img.get('src') if Img else ""
                    h3 = li.find('h3')
                    if h3:
                        Title = h3.find('strong')
                        if Title:
                            Title=Title.text
                        Href = h3.find('a').get('href', "")
                    Time = li.find('span', {'class':'postdate'})
                    Time = Time.text if Time else ""
                    Descript = li.find('p',{'class':'decription'})
                    Descript = Descript.text if Descript else ""
                    print(Time)
                    if len(Title) > 0 and len(Href) > 0 and len(Time) > 0 and len(Descript) > 0 and len(Img) > 0 and topic != 'Hỏi - Đáp':
                        data.append((topic, Img, Title, Href, Time, Descript))
                        # cursor.execute("insert TinTuc values (?,?,?,?,?,?)",(topic,Img,Title,Time,Descript,Href))
                        # cursor.commit()
                        #print(Descript)
                        
        else: 
            print("huhu") 
    return data

def Write_To_Excel(data):
    wb = Workbook() 
    sheet_name='BaiViet'
    worksheet = wb.active
    worksheet.title = sheet_name
    worksheet.append(["Chủ Đề","Ảnh minh họa", "Tiêu đề ", "Link chi tiết", "Thời gian đăng", "Mô tả ngắn"])
    for row in data:
        worksheet.append(row)  
    wb.save('data_TinTuc.xlsx') 
    print("Lưu xong")
        
def read_excel(file_name):
    data = []
    wb = load_workbook(filename=file_name)
    ws = wb.active
    # Bỏ qua dòng đầu tiên (chứa tiêu đề cột)
    rows = ws.iter_rows(min_row=2, values_only=True)
    for row in rows:
        topic,href = row[:2]  # Truy xuất 2 cột đầu tiên trong mỗi dòng
        data.append((topic,href))
    wb.close()
    return data   

if __name__ == '__main__':
    file_name = 'Ds_Chu_De.xlsx'
    data_DH = read_excel(file_name)
    data_post=[]
    z=0
    # for topic,href in data_DH:
    #     url='https://thi.tuyensinh247.com' + href
    #     bg=url[0:len(url)-4-1]
    #     en=url[len(url)-4-1:len(url)]
    #     for i in range(1,11):
    #         finalUrl=bg+"/"+str(i)+en
    #         z+=1
    #         dt = Crawl_News(finalUrl, topic)
    #         #print("Trang: ",i," ",len(data))
    #         data_post.extend(dt)
    # Write_To_Excel(data_post)
    dt = Crawl_News("https://thi.tuyensinh247.com/dap-an-de-thi-c16/1.html", "Test")