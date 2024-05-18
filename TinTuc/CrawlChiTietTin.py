import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import pyodbc
conx = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=MSI\\SQLEXPRESS;DATABASE=Diem_Thi_THPTQG;UID=lekiet;PWD=123456')
cursor = conx.cursor()
def Crawl_CTBV(url, href,title):
    data = []  # Create an empty list to store crawled data
    response = requests.get(url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')    
        div = soup.find('div', {'id': 'mainContent'})
        if div: 
            data.append((title,href, div))
            tmp=str(div)
            cursor.execute("insert ChitietBaiViet values (?,?,?)",(title,href,tmp))
            cursor.commit()
            print(title)
    else:
        print("huhu")
    return data

def Write_to_Excel(data):
    wb = Workbook()
    sheet_name = "ChiTietBaiViet"
    worksheet = wb.active
    worksheet.title = sheet_name
    worksheet.append(["Tiêu đề","Href", "Content"])
    for title,href, div in data:
        worksheet.append([title,href, str(div)])  # Chuyển div thành chuỗi trước khi ghi vào Excel
    wb.save("ChiTietBaiViet.xlsx")
    print('Lưu thành công')


def read_excel(file_name):
    data = []
    wb = load_workbook(filename=file_name)
    ws = wb.active
    rows = ws.iter_rows(min_row=2, values_only=True)
    for row in rows:
        topic, amh, title, href, time, descript = row[:6]
        data.append((topic, amh, title, href, time, descript))
    wb.close()
    return data   

if __name__ == '__main__':
    file_name = 'data_TinTuc.xlsx'
    data_TinTuc = read_excel(file_name)
    data_post_detail = []
    i=0
    for topic, amh, title, href, time, descript in data_TinTuc:
        url = "https://thi.tuyensinh247.com" + href
        i+=1
        print(i)
        da = Crawl_CTBV(url, href,title)
        Write_to_Excel(da)
        #break
    print("Thành công")