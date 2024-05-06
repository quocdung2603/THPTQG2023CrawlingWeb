import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import pyodbc
# Tạo một từ điển để lưu trữ các worksheet theo tên tỉnh
worksheet_dict = {}
diem_thi = {} # Tạo từ điển để lưu trữ thông tin điểm thi của từng bảng
conx = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=MSI\\SQLEXPRESS;DATABASE=Diem_Thi_THPTQG;UID=lekiet;PWD=123456')
cursor = conx.cursor()

def crawl_diemthi(url, wb, sheet_name, tentinh, sbd):
    # Gửi yêu cầu GET để lấy nội dung của trang web
    response = requests.get(url)

    # Kiểm tra xem yêu cầu có thành công không
    if response.status_code == 200:
        # Sử dụng BeautifulSoup để phân tích HTML
        soup = BeautifulSoup(response.content, 'html.parser')

        # Tìm tất cả các bảng dữ liệu trên trang web
        tables = soup.find_all('table', {'role': 'table', 'class': 'table table-striped table-bordered table-hover responsive-table'})

        # Duyệt qua từng bảng để lấy thông tin điểm thi
        for table in tables:
            # Lấy danh sách tên môn học từ hàng đầu tiên của bảng
            mon_hoc = []
            th_tags = table.find_all('th', {'role': 'columnheader'})
            for th in th_tags:
                colspan_value = th.get('colspan')
                if colspan_value and colspan_value == "4":
                    continue
                mon_hoc.append(th.get_text())

            # Lấy dữ liệu điểm từ các hàng trong tbody của bảng
            tbody = table.find('tbody')
            rows = tbody.find_all('tr')
            for row in rows:
                cells = row.find_all('td')
                for i, cell in enumerate(cells):
                    if cell.get_text().strip():
                        mon = mon_hoc[i]
                        diem_thi[mon] = cell.get_text().strip()
                    else:
                        mon = mon_hoc[i]
                        diem_thi[mon] = 0

            # Ghi dữ liệu vào tệp Excel
            print("Số Báo Danh: ",sbd)
            #print(diem_thi)
            cursor.execute("insert Diem_Thi values (?,?,?,?,?,?,?,?,?,?,?,?)",
                        sbd,diem_thi['Toán'],diem_thi['Ngữ văn'],diem_thi['Ngoại ngữ'],diem_thi['Vật lý'],diem_thi['Hóa học'],diem_thi['Sinh học'],diem_thi['Điểm KHTN'],
                        diem_thi['Lịch sử'],diem_thi['Địa lý'],diem_thi['GDCD'],diem_thi['Điểm KHXH'])
            conx.commit()
            #print(diem_thi['Toán'])       
            #write_to_excel(wb, sheet_name, x, diem_thi)

    else:
        print('Failed to retrieve data')

def write_to_excel(wb, sheet_name, sbd, diem_thi):
    # Kiểm tra xem tồn tại sheet_name trong workbook chưa
    if sheet_name not in worksheet_dict:
        worksheet = wb.active
        worksheet.title = sheet_name
        worksheet_dict[sheet_name] = worksheet
        # Chèn một hàng mới vào vị trí thứ 1
        worksheet.insert_rows(1)
        # Ghi số báo danh vào ô A1
        worksheet.cell(row=1, column=1, value="Số báo danh")
        # Ghi tên các môn thi vào các cột tiếp theo dựa trên tên môn học
        for i, (mon, diem) in enumerate(diem_thi.items(), start=2):
            worksheet.cell(row=1, column=i, value=mon)
    else: 
        worksheet = worksheet_dict[sheet_name]
    # Ghi số báo danh vào cột đầu tiên
    worksheet.cell(row=int(sbd)-(startID-1)+1, column=1, value=str(sbd))
    # Đếm số ô dữ liệu bị trống trên 1 dòng
    empty_cnt=0
    # Ghi điểm thi vào các cột tiếp theo dựa trên tên môn học
    for i, (mon, diem) in enumerate(diem_thi.items(), start=2):
        if diem =="": 
            empty_cnt = empty_cnt + 1
        worksheet.cell(row=int(sbd)-(startID-1)+1, column=i, value=diem) 
    if empty_cnt >=11: 
        print("Không có thông tin số báo danh",sbd)
        worksheet.delete_rows(int(sbd)-(startID-1)+1)
        return
    else:
        print("SBD", sbd, "written to Excel, sheet", sheet_name)

def read_tinh_data(file_name):
    data = []
    wb = load_workbook(filename=file_name)
    ws = wb.active

    # Bỏ qua dòng đầu tiên (chứa tiêu đề cột)
    rows = ws.iter_rows(min_row=2, values_only=True)
    
    for row in rows:
        ma_tinh, ten_tinh = row[:2]  # Truy xuất 2 cột đầu tiên trong mỗi dòng
        data.append((ma_tinh, ten_tinh))
    
    wb.close()
    return data

def convert(a, b):
    a = str(a)
    b = str(b)    
    ab = a + b[2:]
    ab = int(ab)
    return ab
    
if __name__ == '__main__':
    file_name = 'data_tinh.xlsx'
    data_tinh = read_tinh_data(file_name)
    for matinh, tentinh in data_tinh:
        wb = Workbook()
        # print(matinh+" "+tentinh)
        # cursor.execute("insert Tinh values (?,?)",(matinh,tentinh))
        # cursor.commit()
        startID = convert(matinh, 11000001)
        endID = convert(matinh, 11000201)
        for x in range(startID, endID, +1):
            if len(str(x)) == 7:
                x = '0' + str(x)
            url = 'https://thptquocgia.edu.vn/diemthi/-/?sbd=' + str(x)
            crawl_diemthi(url, wb, str(tentinh), str(tentinh),str(x))
        tinh = str(tentinh)
        wb.save(f'C:/Users/Dung/Desktop/THPTQG2023/diem_thi_thptqg_2023_{tinh}.xlsx')
        wb.save(f'D:/Project_Code/Python/CuoiKyPhanTichThietKe/diem_thi_thptqg_2023_{tinh}.xlsx')
    conx.close()


#------THPTQG2023--------
#bình dương 44000001 - 440014218
#tphcm 02059988

#wb.save(f'D:/TDMU/Nam3/HK2/KTLTinPTTK/project/diem_thi_thptqg_2023_{tinh}.xlsx')